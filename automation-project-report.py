from openpyxl import load_workbook
from copy import copy
import datetime

# variables for operating the Excel sheet
requiredColumns = [1, 3, 19, 18, 16, 17, 14, 15, 2, 8, 4 ,5, 12, 6, 7, 9, 10, 11]
templatePath = 'progress_report_template.xlsx'
# for report file
cellDate = 'G2'
# project info -> no, name, description, members, start date, deadline, status, progress
projectInfoCells = ['C6', 'C7', 'C8', 'C9', 'C10', 'F10', 'C11', 'F11']
# task info -> no, progress, type, name, details, start date, deadline, planned man-month, actual, ratio
taskInfoCols = [3, 6, 3, 6, 3, 3, 6, 3, 3, 6]
taskInfoRows = [14, 14, 15, 15, 16, 17, 17, 18, 19, 19]
# marge cell column letters
margeStartCols = ['C', 'F', 'C', 'F', 'C', 'C', 'F', 'C', 'C', 'F']
margeEndCols = [':D', ':G', ':D', ':G', ':G', ':D', ':G', ':D', ':D', ':G']

# function to output a progress report with Excel file
def generateProgressReport(filePath):
    # open Execl file and retrieve a target sheet
    wbTaskList = load_workbook(filePath)
    taskSheet = wbTaskList.active
    projects = []
    projectNumbers = []
   
    for rowNum in range(2, 62):
        # create a list and store task information
        taskInfo = []
        for colNum in requiredColumns:
            taskInfo.append(taskSheet.cell(rowNum,colNum).value)

        projectNum = taskInfo[0]
        if projectNum in projectNumbers:
            # append the info to the same index as the same project number
            index = projectNumbers.index(projectNum)
            projects[index].append(taskInfo)
        else:
            # append the info to new index when the project number is difference  
            projects.append([taskInfo]) 
            projectNumbers.append(projectNum)      
    wbTaskList.close()

    # get datetime to use for report date and file name 
    dtNow = datetime.datetime.now()
    currentDate = dtNow.strftime('%m/%d/%Y')
    currentDatetime = dtNow.strftime('%m%d%Y%H%M%S')
    
    # create Excel file by project
    for pNum in projectNumbers:
        index = projectNumbers.index(pNum)
        # open the target excel and sheet
        wbReport = load_workbook(templatePath)
        reportSheet = wbReport.active

        # add date and project info to their proper cell
        reportSheet[cellDate] = currentDate
        for i in range(len(projectInfoCells)):
            reportSheet[projectInfoCells[i]] = projects[index][0][i]

        # get information of tasks
        taskNumbers = []
        count = 0
        for i in range(len(projects[index])):
            if not projects[index][i][8] in taskNumbers:
                taskNumbers.append(projects[index][i][8])
                if count == 0:
                    # add task info to each cell
                    for j in range(len(taskInfoCols)):
                        reportSheet.cell(row=taskInfoRows[j], column=taskInfoCols[j], value=projects[index][i][j+8])
                else:
                    # duplicate task information area before add values if this is not the first task
                    addRows = 7*count
                    for row in range(14, 20):
                        for col in range(2, 8):
                            copyCell = reportSheet.cell(row=row, column=col)
                            pasteCell = reportSheet.cell(row=row+addRows, column=col)
                            pasteCell._style = copy(copyCell._style)
                            pasteCell.value = copyCell.value
                    # modify each cell
                    for j in range(len(taskInfoCols)):
                        newRow = taskInfoRows[j] + addRows
                        if projects[index][i][j+8] == None:
                            reportSheet.cell(row=newRow, column=taskInfoCols[j], value='')
                        else:
                            reportSheet.cell(row=newRow, column=taskInfoCols[j], value=projects[index][i][j+8])
                        # marge cells
                        margeRange = margeStartCols[j] + str(newRow) + margeEndCols[j] + str(newRow)
                        reportSheet.merge_cells(margeRange)
                        if j == 7:
                            # slash(empty) cell area
                            slashCellRange = 'E' + str(newRow) + ':G' + str(newRow)
                            reportSheet.merge_cells(slashCellRange)
                count += 1

        wbReport.save(pNum + '_progress_report_' + currentDatetime +'.xlsx')
        wbReport.close()
    print(projects[0][10])
    return 0

# call the function 
generateProgressReport('task_list.xlsx')
